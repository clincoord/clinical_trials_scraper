import requests
import pandas as pd
import smtplib
from email.message import EmailMessage
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ========== CONFIGURATION ==========

MAX_RESULTS = 1000
SMTP_SERVER = "smtp-mail.outlook.com"
SMTP_PORT = 587
EMAIL_SENDER = "network@clincoord.org"
EMAIL_PASSWORD = "dcpcknbsykkhfcrb"
EMAIL_RECEIVER = [
    "rodrigo.lima.cc.ao@gmail.com"
]
EMAIL_SUBJECT = "New Clinical Trials Found"
XLSX_FILENAME = "clinical_trials_angola.xlsx"

# ========== FLATTEN FUNCTION ==========

def flatten_json(y, prefix=''):
    out = {}
    if isinstance(y, dict):
        for k, v in y.items():
            if k == 'locations' and isinstance(v, list):
                if v:
                    locations_info = []
                    for loc in v:
                        facility = loc.get('facility', 'N/A')
                        city = loc.get('city', 'N/A')
                        state = loc.get('state', None)
                        country = loc.get('country', 'N/A')
                        location_str = f"{facility}, {city}"
                        if state:
                            location_str += f", {state}"
                        location_str += f", {country}"
                        locations_info.append(location_str)
                    out[f"{prefix}combined_locations_string"] = "; ".join(locations_info)
                else:
                    out[f"{prefix}combined_locations_string"] = ""
            elif k == 'centralContacts' and isinstance(v, list):
                if v:
                    first_contact = v[0]
                    out.update(flatten_json(first_contact, f"{prefix}{k}.0."))
            elif k == 'overallOfficials' and isinstance(v, list):
                if v:
                    first_official = v[0]
                    out.update(flatten_json(first_official, f"{prefix}{k}.0."))
            else:
                out.update(flatten_json(v, f"{prefix}{k}."))
    elif isinstance(y, list):
        if all(isinstance(i, (str, int, float, bool)) for i in y):
            out[prefix[:-1]] = ", ".join(map(str, y))
        else:
            for i, item in enumerate(y):
                out.update(flatten_json(item, f"{prefix}{i}."))
    else:
        out[prefix[:-1]] = y
    return out

# ========== SCRAPE FUNCTION ==========

def scrape_clinicaltrials_gov_api(max_results=1000):
    url = "https://clinicaltrials.gov/api/v2/studies"
    params = {"pageSize": max_results}
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers, params=params)

    if response.status_code != 200:
        print(f"Failed to fetch data: HTTP {response.status_code}")
        return []

    try:
        data = response.json()
    except ValueError:
        print("Failed to decode JSON response.")
        return []

    trials = []
    for study in data.get("studies", []):
        flat_study = flatten_json(study)
        trials.append(flat_study)
    return trials

# ========== EXCEL FORMATTING FUNCTION ==========

def format_excel(filename):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 100:
                adjusted_width = 100
            sheet.column_dimensions[column_letter].width = adjusted_width

        sheet.freeze_panes = sheet['A2']
        sheet.auto_filter.ref = sheet.dimensions

        # Torna o link clicável e azul sublinhado
        link_col_idx = None
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == "Trial Link":
                link_col_idx = idx
                break

        if link_col_idx:
            for row in sheet.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx):
                cell = row[0]
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0000FF", underline="single")

        workbook.save(filename)
        print(f"Excel file '{filename}' formatted successfully.")
    except Exception as e:
        print(f"Error formatting Excel file: {e}")

# ========== EMAIL FUNCTION ==========

def send_email(trials_to_send):
    msg = EmailMessage()
    msg['Subject'] = EMAIL_SUBJECT
    msg['From'] = EMAIL_SENDER
    msg['To'] = ", ".join(EMAIL_RECEIVER)

    if trials_to_send:
        html = f"""
        <html>
        <body>
            <h2>Novos Ensaios Clínicos Encontrados</h2>
            <p>{len(trials_to_send)} novo(s) ensaio(s) clínico(s) foi(ram) encontrado(s) e salvo(s) no Excel.</p>
        </body>
        </html>
        """
    else:
        html = """
        <html>
        <body>
            <h2>Nenhum Novo Ensaio Clínico Encontrado</h2>
            <p>Nenhum ensaio clínico correspondente encontrado no momento.</p>
        </body>
        </html>
        """

    msg.set_content("Este email contém conteúdo HTML. Por favor, visualize-o em um cliente de email compatível com HTML.")
    msg.add_alternative(html, subtype='html')

    if trials_to_send:
        try:
            with open(XLSX_FILENAME, "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=XLSX_FILENAME
                )
        except FileNotFoundError:
            print(f"Erro: O arquivo Excel '{XLSX_FILENAME}' não foi encontrado. Não é possível anexá-lo ao email.")
    
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
            smtp.starttls()
            smtp.login(EMAIL_SENDER, EMAIL_PASSWORD)
            smtp.send_message(msg)
            print("Email enviado com sucesso.")
    except smtplib.SMTPAuthenticationError as e:
        print(f"Erro de Autenticação SMTP: {e}. Verifique o remetente e a senha do seu email.")
    except smtplib.SMTPException as e:
        print(f"Erro SMTP: {e}. Não foi possível enviar o email.")
    except Exception as e:
        print(f"Ocorreu um erro inesperado ao enviar o email: {e}")

# ========== MAIN FUNCTION ==========

def main():
    trials = scrape_clinicaltrials_gov_api(MAX_RESULTS)

    if trials:
        df = pd.DataFrame(trials)

        columns_map = {
            "protocolSection.identificationModule.nctId": "Trial Registry Number (.gov)",
            "protocolSection.sponsorCollaboratorsModule.leadSponsor.name": "Sponsor Name",
            "protocolSection.sponsorCollaboratorsModule.leadSponsor.class": "Sponsor Type",
            "protocolSection.contactsLocationsModule.centralContacts.0.name": "Contact Person",
            "protocolSection.contactsLocationsModule.centralContacts.0.role": "Role",
            "protocolSection.contactsLocationsModule.centralContacts.0.phone": "Phone Number",
            "protocolSection.contactsLocationsModule.centralContacts.0.email": "Email",
            "protocolSection.identificationModule.briefTitle": "Trial Name",
            "protocolSection.identificationModule.officialTitle": "Trial/Project Title",
            "protocolSection.designModule.phases": "Trial Phase",
            "protocolSection.statusModule.overallStatus": "Trial Status",
            "protocolSection.conditionsModule.conditions": "Therapeutic Area/Research Category",
            "protocolSection.armsInterventionsModule.interventions.0.name": "Intervention/Investigational Product",
            "protocolSection.statusModule.startDateStruct.date": "Trial Start Date",
            "protocolSection.statusModule.completionDateStruct.date": "Trial End Date",
            "protocolSection.contactsLocationsModule.combined_locations_string": "Location"
        }

        if "protocolSection.identificationModule.nctId" in df.columns:
            df["Trial Link"] = df["protocolSection.identificationModule.nctId"].apply(
                lambda nct: f"https://clinicaltrials.gov/study/{nct}" if pd.notnull(nct) else ""
            )
        else:
            df["Trial Link"] = ""

        df_filtered = pd.DataFrame()
        for api_key, friendly_name in columns_map.items():
            if api_key in df.columns:
                df_filtered[friendly_name] = df[api_key]
            else:
                df_filtered[friendly_name] = ""

        df_filtered["Trial Link"] = df["Trial Link"]

        allowed_statuses = ["RECRUITING", "NOT_YET_RECRUITING"]
        if "Trial Status" in df_filtered.columns:
            df_filtered = df_filtered[df_filtered["Trial Status"].isin(allowed_statuses)].copy()
            print(f"Filtered to {len(df_filtered)} trials with status 'RECRUITING' or 'NOT_YET_RECRUITING'.")
        else:
            print("Warning: 'Trial Status' column not found for filtering.")

        desired_order = list(columns_map.values()) + ["Trial Link"]
        for col in desired_order:
            if col not in df_filtered.columns:
                df_filtered[col] = ""

        df_final = df_filtered  esired_order]

        df_final.to_excel(XLSX_FILENAME, index=False)
        print(f"{len(df_final)} trials saved to '{XLSX_FILENAME}'.")

        format_excel(XLSX_FILENAME)
        send_email(df_final.to_dict('records'))
    else:
        send_email([])
        print("No trials found from API call.")

if __name__ == "__main__":
    main()